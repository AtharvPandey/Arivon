from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from fastapi.responses import Response, FileResponse
from sqlalchemy.orm import Session
import csv
import io
import os
import uuid

from app.database import get_db
from app import models, schemas
from app.core.deps import get_current_user, require_roles

router = APIRouter(prefix="/students", tags=["students"])

MANAGEMENT_ROLES = ("school_admin", "principal", "vice_principal", "administrator", "admissions_officer", "super_admin")


def generate_admission_number(db: Session, school_id: int, academic_year_id: int) -> str:
    """
    Sequential per (school, academic year) — e.g. "2026-0001", "2026-0002".
    Shared between direct student creation and admission enrollment so
    there's exactly one place that knows how this number is built.
    Not atomic under concurrent writes (two Admins enrolling at the
    exact same instant could theoretically collide) — acceptable for a
    single-admin-at-a-time school ERP; the uniqueness check at insert
    time still catches a collision rather than silently duplicating.
    """
    academic_year = db.query(models.AcademicYear).filter(models.AcademicYear.id == academic_year_id).first()
    year_label = academic_year.label.split("-")[0] if academic_year else "0000"

    existing_count = db.query(models.Student).filter(
        models.Student.school_id == school_id, models.Student.academic_year_id == academic_year_id,
    ).count()

    candidate = f"{year_label}-{existing_count + 1:04d}"
    # Guard against a gap-filled sequence colliding (e.g. a manually
    # entered number already took this slot) by nudging forward until free.
    attempt = existing_count + 1
    while db.query(models.Student).filter(
        models.Student.school_id == school_id, models.Student.admission_number == candidate,
    ).first():
        attempt += 1
        candidate = f"{year_label}-{attempt:04d}"
    return candidate


@router.post("/", response_model=schemas.StudentOut, status_code=201)
def create_student(payload: schemas.StudentCreate, db: Session = Depends(get_db)):
    school = db.query(models.School).filter(models.School.id == payload.school_id).first()
    if not school:
        raise HTTPException(status_code=404, detail="School not found")

    data = payload.model_dump()

    if data["admission_number"]:
        # Admission numbers must be unique within a school (not globally —
        # two different schools can both have admission number "1001").
        existing = db.query(models.Student).filter(
            models.Student.school_id == payload.school_id,
            models.Student.admission_number == payload.admission_number,
        ).first()
        if existing:
            raise HTTPException(
                status_code=400,
                detail="Admission number already used in this school",
            )
    else:
        data["admission_number"] = generate_admission_number(db, payload.school_id, payload.academic_year_id)

    student = models.Student(**data)
    db.add(student)
    db.commit()
    db.refresh(student)
    return student


@router.get("/", response_model=list[schemas.StudentOut])
def list_students(
    school_id: int,
    section_id: int | None = None,
    house_id: int | None = None,
    category: str | None = None,
    gender: str | None = None,
    search: str | None = None,
    include_inactive: bool = False,
    db: Session = Depends(get_db),
):
    """
    Extended with the Student Directory's real filters — search by
    name/admission number, plus category/gender/house — none of which
    existed before; this endpoint only ever supported a bare section_id
    filter.
    """
    query = db.query(models.Student).filter(models.Student.school_id == school_id)
    if not include_inactive:
        query = query.filter(models.Student.is_active == True)  # noqa: E712
    if section_id is not None:
        query = query.filter(models.Student.section_id == section_id)
    if house_id is not None:
        query = query.filter(models.Student.house_id == house_id)
    if category:
        query = query.filter(models.Student.category == category)
    if gender:
        query = query.filter(models.Student.gender == gender)
    if search:
        like = f"%{search}%"
        query = query.filter(
            (models.Student.full_name.ilike(like)) | (models.Student.admission_number.ilike(like))
        )
    return query.all()


@router.get("/export")
def export_students(
    school_id: int,
    section_id: int | None = None,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    """
    CSV export — per the School Admin plan, this is the single
    most-requested feature in tier 2/3 schools, since boards and
    government portals routinely require student lists in specific
    formats. Placed BEFORE /{student_id} in this file so FastAPI
    doesn't try to parse "export" as a student ID.
    """
    students = _export_query(db, school_id, section_id)

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(EXPORT_HEADERS)
    for s in students:
        writer.writerow(_export_row(s))

    return Response(
        content=output.getvalue(), media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=students_export.csv"},
    )


EXPORT_HEADERS = [
    "Admission Number", "Full Name", "Date of Birth", "Gender", "Blood Group",
    "Category", "Religion", "Nationality", "Mother Tongue", "Guardian Name",
    "Guardian Phone", "Guardian Email", "Aadhaar Number", "Previous School",
]


def _export_query(db: Session, school_id: int, section_id: int | None):
    query = db.query(models.Student).filter(
        models.Student.school_id == school_id, models.Student.is_active == True,  # noqa: E712
    )
    if section_id is not None:
        query = query.filter(models.Student.section_id == section_id)
    return query.order_by(models.Student.admission_number).all()


def _export_row(s: models.Student) -> list:
    return [
        s.admission_number, s.full_name, s.date_of_birth.isoformat(), s.gender or "",
        s.blood_group or "", s.category or "", s.religion or "", s.nationality or "",
        s.mother_tongue or "", s.guardian_name, s.guardian_phone, s.guardian_email or "",
        s.aadhaar_number or "", s.previous_school or "",
    ]


@router.get("/export.xlsx")
def export_students_excel(
    school_id: int,
    section_id: int | None = None,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    """
    The actual Excel version — CSV alone doesn't give schools what board
    and government portals often expect (a real spreadsheet, a bold
    header row, sensible column widths). Same data as the CSV export,
    same shared query, just written with openpyxl instead of the csv module.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    students = _export_query(db, school_id, section_id)

    wb = Workbook()
    ws = wb.active
    ws.title = "Students"

    header_font = Font(name="Calibri", bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="6D5BFF", end_color="6D5BFF", fill_type="solid")
    body_font = Font(name="Calibri")

    for col_idx, header in enumerate(EXPORT_HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for row_idx, s in enumerate(students, start=2):
        for col_idx, value in enumerate(_export_row(s), start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = body_font

    # Column widths sized to the header text, with a sensible floor —
    # good enough without inspecting every cell's actual content length.
    for col_idx, header in enumerate(EXPORT_HEADERS, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = max(len(header) + 4, 14)

    ws.freeze_panes = "A2"

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return Response(
        content=buffer.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=students_export.xlsx"},
    )


@router.post(
    "/bulk/section-shuffle", response_model=schemas.BulkOperationResult,
    dependencies=[Depends(require_roles(*MANAGEMENT_ROLES))],
)
def bulk_section_shuffle(payload: schemas.BulkSectionShuffleRequest, db: Session = Depends(get_db)):
    """Redistribute a list of students into a different section — e.g.
    balancing class strength across sections at the start of a term."""
    new_section = db.query(models.Section).filter(models.Section.id == payload.new_section_id).first()
    if not new_section:
        raise HTTPException(status_code=404, detail="Target section not found")

    succeeded, failed = [], []
    for student_id in payload.student_ids:
        student = db.query(models.Student).filter(models.Student.id == student_id).first()
        if not student:
            failed.append({"student_id": student_id, "reason": "not found"})
            continue
        student.section_id = payload.new_section_id
        succeeded.append(student_id)

    db.commit()
    return schemas.BulkOperationResult(succeeded=succeeded, failed=failed)


@router.post(
    "/bulk/promote", response_model=schemas.BulkOperationResult,
    dependencies=[Depends(require_roles(*MANAGEMENT_ROLES))],
)
def bulk_promote(payload: schemas.BulkPromoteRequest, db: Session = Depends(get_db)):
    """Promotes every active student in one section into a section in
    the next academic year — the real end-of-year workflow, not a
    per-student manual edit repeated dozens of times."""
    target_section = db.query(models.Section).filter(models.Section.id == payload.target_section_id).first()
    if not target_section:
        raise HTTPException(status_code=404, detail="Target section not found")
    target_year = db.query(models.AcademicYear).filter(models.AcademicYear.id == payload.target_academic_year_id).first()
    if not target_year:
        raise HTTPException(status_code=404, detail="Target academic year not found")

    students = db.query(models.Student).filter(
        models.Student.section_id == payload.source_section_id, models.Student.is_active == True,  # noqa: E712
    ).all()

    succeeded, failed = [], []
    for student in students:
        student.section_id = payload.target_section_id
        student.academic_year_id = payload.target_academic_year_id
        succeeded.append(student.id)

    db.commit()
    return schemas.BulkOperationResult(succeeded=succeeded, failed=failed)


@router.get("/{student_id}", response_model=schemas.StudentOut)
def get_student(student_id: int, db: Session = Depends(get_db)):
    student = db.query(models.Student).filter(models.Student.id == student_id).first()
    if not student:
        raise HTTPException(status_code=404, detail="Student not found")
    return student


@router.get("/{student_id}/detail", response_model=schemas.StudentDetail)
def get_student_detail(student_id: int, db: Session = Depends(get_db)):
    """
    Fuller record for the "click a student, see everything" drill-down —
    used by the Students tab. Bank details are deliberately NOT included
    here; those stay behind the Finance-only /fees/students/{id}/bank-details
    endpoint regardless of who's viewing this page.
    """
    student = db.query(models.Student).filter(models.Student.id == student_id).first()
    if not student:
        raise HTTPException(status_code=404, detail="Student not found")
    return student


@router.patch(
    "/{student_id}", response_model=schemas.StudentDetail,
    dependencies=[Depends(require_roles(*MANAGEMENT_ROLES))],
)
def update_student(student_id: int, payload: schemas.StudentUpdate, db: Session = Depends(get_db)):
    """
    The piece that was missing entirely: editing a student's profile
    after creation. Genuinely partial — only fields present in the
    request body get touched (model_dump(exclude_unset=True)), so a
    request that only wants to update medical_notes doesn't
    accidentally null out every other field.
    """
    student = db.query(models.Student).filter(models.Student.id == student_id).first()
    if not student:
        raise HTTPException(status_code=404, detail="Student not found")

    updates = payload.model_dump(exclude_unset=True)
    for field, value in updates.items():
        setattr(student, field, value)

    db.commit()
    db.refresh(student)
    return student


@router.post(
    "/{student_id}/photo", response_model=schemas.StudentOut,
    dependencies=[Depends(require_roles(*MANAGEMENT_ROLES))],
)
async def upload_student_photo(student_id: int, file: UploadFile = File(...), db: Session = Depends(get_db)):
    """
    A dedicated, simple endpoint for the one specific case of a profile
    photo — deliberately NOT routed through the generic polymorphic
    Documents system (which is right for birth certificates/TCs/Aadhaar,
    but would be a heavier round-trip than necessary for "just set this
    student's photo"). Validates it's actually an image and reasonably
    sized before accepting it.
    """
    ALLOWED_IMAGE_EXTENSIONS = {".jpg", ".jpeg", ".png"}
    MAX_PHOTO_SIZE_BYTES = 3 * 1024 * 1024  # 3MB — a profile photo, not a scanned document

    student = db.query(models.Student).filter(models.Student.id == student_id).first()
    if not student:
        raise HTTPException(status_code=404, detail="Student not found")

    ext = os.path.splitext(file.filename)[1].lower()
    if ext not in ALLOWED_IMAGE_EXTENSIONS:
        raise HTTPException(status_code=400, detail="Photo must be a JPG or PNG file")

    contents = await file.read()
    if len(contents) > MAX_PHOTO_SIZE_BYTES:
        raise HTTPException(status_code=400, detail=f"Photo is too large ({len(contents) / 1024 / 1024:.1f}MB). Maximum size is 3MB.")

    photo_dir = "uploads/photos"
    os.makedirs(photo_dir, exist_ok=True)
    stored_filename = f"{uuid.uuid4().hex}{ext}"
    with open(os.path.join(photo_dir, stored_filename), "wb") as f:
        f.write(contents)

    student.photo_url = f"/uploads/photos/{stored_filename}"
    db.commit()
    db.refresh(student)
    return student


@router.get("/{student_id}/siblings", response_model=list[schemas.SiblingItem])
def get_siblings(student_id: int, db: Session = Depends(get_db)):
    """
    Other children of the same guardian — structurally this has always
    been possible (Student.guardian_id already links siblings together),
    it just never had a query surfacing it. Very common in Indian
    schools; this is what makes "sibling linking" a visible feature
    rather than an unused foreign key.
    """
    student = db.query(models.Student).filter(models.Student.id == student_id).first()
    if not student:
        raise HTTPException(status_code=404, detail="Student not found")
    if not student.guardian_id:
        return []

    siblings = db.query(models.Student).filter(
        models.Student.guardian_id == student.guardian_id, models.Student.id != student_id,
    ).all()

    result = []
    for sibling in siblings:
        section = db.query(models.Section).filter(models.Section.id == sibling.section_id).first() if sibling.section_id else None
        school_class = db.query(models.SchoolClass).filter(models.SchoolClass.id == section.school_class_id).first() if section else None
        section_name = f"{school_class.name} - {section.name}" if section and school_class else None
        result.append(schemas.SiblingItem(
            id=sibling.id, full_name=sibling.full_name,
            admission_number=sibling.admission_number, section_name=section_name,
        ))
    return result


@router.post(
    "/{student_id}/readmit", response_model=schemas.StudentOut,
    dependencies=[Depends(require_roles(*MANAGEMENT_ROLES))],
)
def readmit_student(student_id: int, payload: schemas.ReadmitRequest, db: Session = Depends(get_db)):
    """Reactivates a student who left and has now returned — common
    enough in Indian schools (family relocated and came back, financial
    gap year, etc.) to be its own explicit action rather than a manual
    is_active flip with no audit trail of what happened."""
    student = db.query(models.Student).filter(models.Student.id == student_id).first()
    if not student:
        raise HTTPException(status_code=404, detail="Student not found")
    if student.is_active:
        raise HTTPException(status_code=400, detail="This student is already active")

    student.is_active = True
    student.academic_year_id = payload.academic_year_id
    student.section_id = payload.section_id
    student.date_of_leaving = None
    student.leaving_reason = None
    db.commit()
    db.refresh(student)
    return student


@router.post(
    "/{student_id}/transfer-certificate", response_model=schemas.GenerateTCResponse,
    dependencies=[Depends(require_roles(*MANAGEMENT_ROLES))],
)
def generate_transfer_certificate(
    student_id: int,
    payload: schemas.GenerateTCRequest,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    """
    Generates the actual PDF, stores it as a Document (reusing the
    existing polymorphic Document model — no separate certificates
    table), marks the student inactive with a leaving date/reason, and
    stamps a permanent TC number that's never reused even if this
    endpoint is somehow called again for the same student.
    """
    from app.core.certificates import generate_transfer_certificate_pdf

    student = db.query(models.Student).filter(models.Student.id == student_id).first()
    if not student:
        raise HTTPException(status_code=404, detail="Student not found")
    if student.tc_number:
        raise HTTPException(status_code=400, detail=f"A Transfer Certificate ({student.tc_number}) already exists for this student")

    school = db.query(models.School).filter(models.School.id == student.school_id).first()
    section = db.query(models.Section).filter(models.Section.id == student.section_id).first() if student.section_id else None
    school_class = db.query(models.SchoolClass).filter(models.SchoolClass.id == section.school_class_id).first() if section else None

    year_label = str(payload.date_of_leaving.year)
    tc_count = db.query(models.Student).filter(
        models.Student.school_id == student.school_id, models.Student.tc_number.isnot(None),
    ).count()
    tc_number = f"TC-{year_label}-{tc_count + 1:04d}"

    student.date_of_leaving = payload.date_of_leaving
    student.leaving_reason = payload.leaving_reason
    student.tc_number = tc_number
    student.is_active = False

    stored_filename = generate_transfer_certificate_pdf(
        student, school, school_class.name if school_class else "—",
        section.name if section else None, tc_number,
    )

    document = models.Document(
        school_id=student.school_id, entity_type="student", entity_id=student.id,
        document_type="transfer_certificate", original_filename=f"TC_{student.full_name.replace(' ', '_')}.pdf",
        stored_filename=stored_filename, uploaded_by_user_id=current_user.id,
    )
    db.add(document)
    db.commit()
    db.refresh(document)

    return schemas.GenerateTCResponse(
        document_id=document.id, tc_number=tc_number,
        download_url=f"/documents/{document.id}/download",
    )
